import hashlib
import os
import re
import zipfile
from pathlib import Path

import openpyxl #not sure if I need this - it writes excel files
from pipeline import exceptions


def get_analysis_engine(analysis_request):
    engine = analysis_request["metadata"]["engine"]
    engine = re.sub("-.*", "", engine)
    engine = engine.lower()
    return engine


def get_request_type(analysis_request):
    request_id = analysis_request["metadata"]["id"]
    req_type = re.sub("_0000", "", request_id)
    req_type = re.sub(r".+?\_", "", req_type)
    return req_type


def get_request_sha(analysis_request):
    request_id = analysis_request["metadata"]["id"]
    hash_ = hashlib.sha1(request_id.encode("utf-8"))
    return hash_.hexdigest()


def get_parent_dir(file_path: str) -> str:
    _file_path = Path(file_path)
    return _file_path.parent


def create_workbook(workbook_file: str, sheet_names: list[str]):
    """Creates a excel workbook for given file path.

    Args:
        workbook_file: path of the excel file to be created.
        sheet_names: initializes the workbook with given sheet names.
    """

    wb = openpyxl.workbook.Workbook()

    for sheet in sheet_names:
        wb.create_sheet(sheet)

    wb.save(filename=workbook_file)


def remove_empty_worksheets(workbook_file: str):
    """Removes empty worksheets in given workbook

    Args:
        workbook_file: path to the excel file

    """

    if not os.path.isfile(workbook_file):
        raise exceptions.InvalidFilePath("Workbook file not found.")

    wb = openpyxl.load_workbook(workbook_file)

    for sheet_name in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(sheet_name)

        if sheet.max_row == 1 and sheet.max_column == 1:
            wb.remove_sheet(sheet)

    wb.save(workbook_file)


def zip_dir(dir_to_zip: str, zip_file_path: str, base_dir: str):
    """Zips the input directory to destination zip file"""

    files_to_zip = []

    for root, directories, files in os.walk(dir_to_zip):

        for file_name in files:

            file_path = os.path.join(root, file_name)
            files_to_zip.append(file_path)

    _zip_file = zipfile.ZipFile(zip_file_path, "a")

    with _zip_file:
        for _file in files_to_zip:
            __zip_file__(_file, _zip_file, base_dir)


def zip_file(file_to_zip: str, zip_file_path: str, base_dir: str = ""):

    _zip_file = zipfile.ZipFile(zip_file_path, "a")

    with _zip_file:
        __zip_file__(file_to_zip, _zip_file, base_dir)


def __zip_file__(file_to_zip: str, zip_file: zipfile.ZipFile, base_dir: str = ""):

    _dest_name = f"{base_dir}/{os.path.basename(file_to_zip)}"
    zip_file.write(file_to_zip, _dest_name, zipfile.ZIP_DEFLATED)
